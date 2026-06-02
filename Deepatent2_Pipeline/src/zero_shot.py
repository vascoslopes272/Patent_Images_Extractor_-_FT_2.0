"""
zero_shot.py — DINOv2 zero-shot embedding extraction, clustering, and visualization.

All heavy logic for the DINOv2_untrained_VTOL_Analysis notebook lives here.
The notebook only configures paths/hyperparameters and calls these functions.

PUBLIC API
──────────
    collect_image_paths(image_dir)                          → list[Path]
    patent_id_from_path(path)                               → str
    category_from_path(path)                                → str
    initialize_dinov2(model_name, device)                    → (processor, model)
    extract_embeddings(image_paths, processor, model,
                       device, batch_size)                  → (image_emb, img_meta_df,
                                                               patent_ids, patent_emb)
    l2_normalize(embeddings)                                → np.ndarray
    pca_reduce(X, n_components, seed)                       → np.ndarray
    hdbscan_cluster(X_pca, patent_ids,
                    min_cluster_size, min_samples)          → (cluster_df, labels)
    dbscan_cluster(X_norm, patent_ids, min_samples,
                   eps_candidates)                          → (cluster_df, labels, best_eps)
    umap_project(X, seed)                                   → np.ndarray
    plot_umap_clusters(points_2d, labels, patent_ids,
                       title_suffix, ground_truth_labels)   → matplotlib.Figure
    plot_cluster_gallery(cluster_df, patent_ids,
                         image_paths, plot_dir)             → None  (saves + shows)

    # ── DeepPatent2 analysis additions ──
    plot_knn_neighbors(query_idx, image_paths,
                       image_emb, k)                        → matplotlib.Figure
    plot_reduction_comparison(patent_emb, patent_ids,
                              hdbscan_min_cluster_size,
                              hdbscan_min_samples,
                              pca_n_components, seed)       → matplotlib.Figure
    export_analysis_matrix(image_paths, dataset_df,
                           cluster_df, output_path)         → pd.DataFrame

    safe_save_np(array, path)   → Path
    safe_save_df(df, path)      → Path
    safe_save_plot(fig, path)   → Path
"""

import time
import warnings
from collections import defaultdict
from pathlib import Path

import hdbscan as hdbscan_lib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import torch
import umap as umap_lib
from PIL import Image
from sklearn.cluster import DBSCAN
from sklearn.decomposition import PCA
from sklearn.metrics import davies_bouldin_score, silhouette_score
from sklearn.preprocessing import normalize
from torch.utils.data import DataLoader, Dataset
from tqdm import tqdm
from transformers import AutoImageProcessor, AutoModel


# ── Constants ─────────────────────────────────────────────────────────────────

VALID_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
_DEFAULT_EPS_CANDIDATES = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]


# ── File-system helpers ───────────────────────────────────────────────────────

def collect_image_paths(image_dir: Path) -> list[Path]:
    """Recursively collect all valid image files under image_dir, sorted."""
    image_dir = Path(image_dir)
    if not image_dir.exists():
        print(f"⚠️  Directory not found: {image_dir}")
        return []
    paths = [
        p for p in image_dir.rglob("*")
        if p.is_file() and p.suffix.lower() in VALID_EXTENSIONS
    ]
    return sorted(paths)


def patent_id_from_path(path: Path) -> str:
    """Extract patent ID from filename stem (e.g. US2022267016A1_SHR_... → US2022267016A1)."""
    return path.name.split("_")[0]


def category_from_path(path: Path) -> str:
    """Extract class label from filename: *_SHR_* → 'shrouded', *_OPN_* → 'open_rotor'."""
    name = path.name.upper()
    if "_SHR_" in name:
        return "shrouded"
    if "_OPN_" in name:
        return "open_rotor"
    return "unknown"


def _unique_path(path: Path) -> Path:
    """Return path unchanged if it does not exist; otherwise append a timestamp."""
    if not path.exists():
        return path
    ts = int(time.time())
    return path.with_name(f"{path.stem}_{ts}{path.suffix}")


def safe_save_np(array: np.ndarray, path: Path) -> Path:
    path = _unique_path(Path(path))
    np.save(path, array)
    return path


def safe_save_df(df: pd.DataFrame, path: Path, **kwargs) -> Path:
    path = _unique_path(Path(path))
    df.to_csv(path, index=kwargs.get("index", False))
    return path


def safe_save_plot(fig: plt.Figure, path: Path, **kwargs) -> Path:
    path = _unique_path(Path(path))
    fig.savefig(path, **kwargs)
    return path


# ── DINOv2 extraction ─────────────────────────────────────────────────────────

class _ImageFolderDataset(Dataset):
    """Internal dataset: returns (preprocessed_tensor, path_stem) per image."""

    def __init__(self, image_paths: list[Path], processor):
        self.image_paths = image_paths
        self.processor   = processor

    def __len__(self):
        return len(self.image_paths)

    def __getitem__(self, idx):
        path  = self.image_paths[idx]
        image = Image.open(path).convert("RGB")
        inputs = self.processor(images=image, return_tensors="pt")
        return inputs["pixel_values"].squeeze(0), path.stem


def _collate_fn(batch):
    return torch.stack([b[0] for b in batch]), [b[1] for b in batch]


def initialize_dinov2(model_name: str, device: torch.device):
    """Load DINOv2 processor + model, move to device, and set eval mode."""
    processor = AutoImageProcessor.from_pretrained(model_name)
    model     = AutoModel.from_pretrained(model_name)
    model.to(device).eval()
    return processor, model


def extract_embeddings(
    image_paths: list[Path],
    processor,
    model,
    device: torch.device,
    batch_size: int = 16,
) -> tuple[np.ndarray, pd.DataFrame, list[str], np.ndarray]:
    """
    Extract DINOv2 CLS-token embeddings for every image, then mean-pool per patent.

    Parameters
    ----------
    image_paths : list of image Paths
    processor   : HuggingFace AutoImageProcessor
    model       : HuggingFace AutoModel (DINOv2)
    device      : torch.device
    batch_size  : DataLoader batch size

    Returns
    -------
    image_emb    : (N_images, D)  per-image CLS vectors
    img_meta_df  : DataFrame [image_path, patent_id]
    patent_ids   : sorted list of unique patent IDs
    patent_emb   : (N_patents, D) mean-pooled patent vectors
    """
    loader = DataLoader(
        _ImageFolderDataset(image_paths, processor),
        batch_size=batch_size,
        shuffle=False,
        num_workers=4,
        collate_fn=_collate_fn,
    )

    all_embeddings: list[np.ndarray] = []
    all_names: list[str] = []

    with torch.no_grad():
        for pixel_values, names in tqdm(loader, desc="Extracting embeddings"):
            cls_tokens = (
                model(pixel_values=pixel_values.to(device))
                .last_hidden_state[:, 0, :]
                .cpu().numpy()
            )
            all_embeddings.append(cls_tokens)
            all_names.extend(names)

    image_emb    = np.vstack(all_embeddings)
    stem_to_path = {p.stem: p for p in image_paths}
    file_paths   = [stem_to_path[n] for n in all_names]
    pids_per_img = [patent_id_from_path(p) for p in file_paths]

    img_meta_df = pd.DataFrame({
        "image_path": [str(p) for p in file_paths],
        "patent_id":  pids_per_img,
    })

    patent_to_vecs: dict[str, list] = defaultdict(list)
    for vec, pid in zip(image_emb, pids_per_img):
        patent_to_vecs[pid].append(vec)

    patent_ids = sorted(patent_to_vecs.keys())
    patent_emb = np.array([np.mean(patent_to_vecs[pid], axis=0) for pid in patent_ids])

    print(f"Image embeddings:  {image_emb.shape}")
    print(f"Patent embeddings: {patent_emb.shape}  ({len(patent_ids)} unique patents)")
    return image_emb, img_meta_df, patent_ids, patent_emb


# ── Normalization ─────────────────────────────────────────────────────────────

def l2_normalize(embeddings: np.ndarray) -> np.ndarray:
    """
    Project embedding rows onto the unit hypersphere (L2 norm = 1).

    After normalization Euclidean(u, v) = sqrt(2 − 2·cos θ), which is a
    monotone function of cosine distance.  DBSCAN with Euclidean metric
    therefore clusters by angular (structural) similarity.
    """
    X_norm = normalize(embeddings, norm="l2")
    norms  = np.linalg.norm(X_norm, axis=1)
    print(f"L2-normalized {X_norm.shape[0]} vectors — "
          f"mean norm: {norms.mean():.6f}, std: {norms.std():.2e}")
    return X_norm


# ── PCA dimensionality reduction ─────────────────────────────────────────────

def pca_reduce(X: np.ndarray, n_components: int = 100, seed: int = 42) -> np.ndarray:
    """
    Reduce embedding matrix from D dimensions to n_components using PCA.

    WHY PCA BEFORE CLUSTERING?
    DINOv2-large produces 1024-d vectors. In very high dimensions all pairwise
    distances converge to the same value (curse of dimensionality), making
    distance-based clustering unreliable. PCA projects onto the n_components
    directions of maximum variance, preserving the most discriminative structure
    while making distances meaningful again.

    Parameters
    ----------
    X            : (N, D) L2-normalised embedding matrix
    n_components : target dimension (your advisor's recommendation: 100)

    Returns
    -------
    X_pca : (N, n_components) reduced matrix
    """
    n_components = min(n_components, X.shape[0], X.shape[1])
    pca = PCA(n_components=n_components, random_state=seed)
    X_pca = pca.fit_transform(X)
    var_explained = pca.explained_variance_ratio_.sum() * 100
    print(f"PCA: {X.shape[1]}d → {n_components}d  |  "
          f"variance explained: {var_explained:.1f}%")
    return X_pca


# ── HDBSCAN clustering ────────────────────────────────────────────────────────

def hdbscan_cluster(
    X_pca: np.ndarray,
    patent_ids: list[str],
    min_cluster_size: int = 5,
    min_samples: int | None = None,
    cluster_selection_method: str = "leaf",
) -> tuple[pd.DataFrame, np.ndarray]:
    """
    Cluster PCA-reduced embeddings with HDBSCAN.

    WHY HDBSCAN INSTEAD OF DBSCAN?
    DBSCAN requires tuning `eps` (the neighbourhood radius), which is hard to
    set correctly — especially with only 84 points. HDBSCAN builds a hierarchy
    of density-connected clusters and extracts the most stable ones automatically.
    It only needs `min_cluster_size` (minimum meaningful group size), which is
    intuitive: set to ~5% of your dataset (84 × 0.05 ≈ 4–5).

    Parameters
    ----------
    X_pca                    : (N, d) PCA-reduced, L2-normalised embeddings
    patent_ids               : list of patent IDs aligned with X_pca rows
    min_cluster_size         : smallest cluster HDBSCAN will recognise (try 4–8 for 84 images)
    min_samples              : controls noise sensitivity; lower = fewer noise points
    cluster_selection_method : "leaf" (tight clusters, good for binary) or "eom" (default HDBSCAN)

    Returns
    -------
    cluster_df : DataFrame [patent_id, cluster_id, cluster_prob]
    labels     : (N,) int array (-1 = noise/unclustered)
    """
    clusterer = hdbscan_lib.HDBSCAN(
        min_cluster_size=min_cluster_size,
        min_samples=min_samples,
        metric="euclidean",
        cluster_selection_method=cluster_selection_method,
    )
    labels = clusterer.fit_predict(X_pca)
    probs  = clusterer.probabilities_

    n_clusters = len(set(labels) - {-1})
    n_noise    = int(np.sum(labels == -1))
    print(f"HDBSCAN (min_cluster_size={min_cluster_size}, min_samples={min_samples}, "
          f"method={cluster_selection_method}): "
          f"{n_clusters} cluster(s), {n_noise} noise point(s) / {len(labels)} total")

    if n_clusters >= 2:
        mask = labels != -1
        sil  = silhouette_score(X_pca[mask], labels[mask])
        db   = davies_bouldin_score(X_pca[mask], labels[mask])
        print(f"Silhouette score: {sil:.4f}  |  Davies-Bouldin: {db:.4f}")

    cluster_df = (
        pd.DataFrame({
            "patent_id":    patent_ids,
            "cluster_id":   labels.astype(int),
            "cluster_prob": np.round(probs, 4),
        })
        .sort_values(["cluster_id", "patent_id"])
        .reset_index(drop=True)
    )
    return cluster_df, labels


# ── DBSCAN clustering ─────────────────────────────────────────────────────────

def dbscan_cluster(
    X_norm: np.ndarray,
    patent_ids: list[str],
    min_samples: int = 3,
    eps_candidates: list[float] | None = None,
) -> tuple[pd.DataFrame, np.ndarray, float]:
    """
    Sweep eps values and select the best DBSCAN configuration by Silhouette Score.

    Noise points (label = −1) are retained in the returned DataFrame.
    Silhouette and Davies-Bouldin are computed on non-noise points only.

    Returns
    -------
    cluster_df            : DataFrame [patent_id, cluster_id], sorted
    patent_cluster_labels : (N_patents,) int array
    effective_eps         : eps that maximised Silhouette Score
    """
    if eps_candidates is None:
        eps_candidates = _DEFAULT_EPS_CANDIDATES

    best_labels     = None
    best_eps        = eps_candidates[-1]
    best_silhouette = -2.0
    results         = []

    for eps in eps_candidates:
        labels  = DBSCAN(eps=eps, min_samples=min_samples, metric="euclidean").fit_predict(X_norm)
        n_clust = len(set(labels) - {-1})
        n_noise = int(np.sum(labels == -1))

        if n_clust >= 2:
            mask = labels != -1
            sil  = silhouette_score(X_norm[mask], labels[mask])
            db   = davies_bouldin_score(X_norm[mask], labels[mask])
        else:
            sil, db = -1.0, float("inf")

        results.append({"eps": eps, "n_clusters": n_clust, "n_noise": n_noise,
                        "silhouette": sil, "davies_bouldin": db})

        if sil > best_silhouette and n_clust >= 2:
            best_silhouette, best_eps, best_labels = sil, eps, labels.copy()

    # ── grid-search table ─────────────────────────────────────────────────
    print("=" * 70)
    hdr = f"{'eps':>5}  {'clusters':>8}  {'noise':>6}  {'silhouette':>11}  {'davies-bouldin':>14}"
    print(hdr)
    print("-" * len(hdr))
    for r in results:
        mark    = " ✓" if r["eps"] == best_eps else "  "
        sil_str = f"{r['silhouette']:.4f}"     if r["silhouette"] > -1.5       else "     N/A"
        db_str  = f"{r['davies_bouldin']:.4f}" if r["davies_bouldin"] < float("inf") else "     N/A"
        print(f"{r['eps']:>5.1f}  {r['n_clusters']:>8}  {r['n_noise']:>6}  "
              f"{sil_str:>11}  {db_str:>14}{mark}")
    print("=" * 70)

    if best_labels is None:
        warnings.warn("No valid DBSCAN clustering found. All points assigned as noise. "
                      "Try reducing min_samples or widening eps_candidates.")
        best_labels = np.full(len(patent_ids), -1, dtype=int)
    else:
        n_final = len(set(best_labels) - {-1})
        n_noise = int(np.sum(best_labels == -1))
        print(f"Selected eps={best_eps:.1f}  →  {n_final} cluster(s), {n_noise} noise point(s)")
        print(f"Best Silhouette: {best_silhouette:.4f}")

    cluster_df = (
        pd.DataFrame({"patent_id": patent_ids, "cluster_id": best_labels.astype(int)})
        .sort_values(["cluster_id", "patent_id"])
        .reset_index(drop=True)
    )
    return cluster_df, best_labels, best_eps


# ── UMAP projection ───────────────────────────────────────────────────────────

def umap_project(X_norm: np.ndarray, seed: int = 42) -> np.ndarray:
    """
    Reduce L2-normalized embeddings to 2D with UMAP.

    UMAP preserves local neighbourhood topology better than PCA in high-d space,
    making cluster separations visible in the 2D scatter plot.
    """
    print("Computing UMAP 2D projection (may take ~30 s on CPU)...")
    points_2d = umap_lib.UMAP(
        n_components=2, n_neighbors=15, min_dist=0.1,
        metric="euclidean", random_state=seed,
    ).fit_transform(X_norm)
    print(f"✓ UMAP complete: {points_2d.shape}")
    return points_2d


# ── Plotting ──────────────────────────────────────────────────────────────────

def plot_umap_clusters(
    points_2d: np.ndarray,
    patent_cluster_labels: np.ndarray,
    patent_ids: list[str],
    title_suffix: str = "",
    ground_truth_labels: list[str] | None = None,
) -> plt.Figure:
    """
    Seaborn scatter of the UMAP 2D projection coloured by HDBSCAN cluster label.
    Noise points (label = −1) are shown in grey.

    When ground_truth_labels is provided (list of "shrouded" / "open_rotor" strings
    aligned with patent_ids), a second panel is shown side-by-side so you can
    directly compare cluster assignments against the true class labels.

    Returns the Figure; the caller saves it with safe_save_plot().
    """
    n_panels = 2 if ground_truth_labels is not None else 1
    fig, axes = plt.subplots(1, n_panels, figsize=(11 * n_panels, 8))
    if n_panels == 1:
        axes = [axes]

    # ── left panel: HDBSCAN cluster assignments ───────────────────────────
    ax = axes[0]
    unique_labels = sorted(set(patent_cluster_labels))
    n_real        = sum(1 for l in unique_labels if l >= 0)
    palette       = sns.color_palette("tab10", n_colors=max(n_real, 1))

    color_map, ci = {}, 0
    for lbl in unique_labels:
        color_map[lbl] = (0.55, 0.55, 0.55) if lbl == -1 else palette[ci]
        if lbl >= 0:
            ci += 1

    labels_str  = [f"Cluster {l}" if l >= 0 else "Noise" for l in patent_cluster_labels]
    str_palette = {
        (f"Cluster {l}" if l >= 0 else "Noise"): color_map[l]
        for l in unique_labels
    }

    plot_df = pd.DataFrame({
        "UMAP-1":    points_2d[:, 0],
        "UMAP-2":    points_2d[:, 1],
        "Cluster":   labels_str,
        "Patent ID": patent_ids,
    })

    sns.scatterplot(
        data=plot_df, x="UMAP-1", y="UMAP-2", hue="Cluster",
        palette=str_palette, s=90, alpha=0.85,
        edgecolor="black", linewidth=0.4, ax=ax,
    )
    suffix = f" — {title_suffix}" if title_suffix else ""
    ax.set_title(
        f"HDBSCAN Cluster Assignments{suffix}\n"
        f"(DINOv2-Large → L2 norm → PCA → HDBSCAN)",
        fontsize=12, fontweight="bold",
    )
    ax.set_xlabel("UMAP Dimension 1", fontsize=11)
    ax.set_ylabel("UMAP Dimension 2", fontsize=11)
    ax.legend(title="Cluster", bbox_to_anchor=(1.02, 1), loc="upper left", framealpha=0.9)
    ax.grid(alpha=0.3)

    # ── right panel: ground-truth class labels ────────────────────────────
    if ground_truth_labels is not None:
        ax2 = axes[1]
        gt_categories = sorted(set(ground_truth_labels))
        # Default palette covers both the old VTOL experiment labels and
        # the DeepPatent2 platform labels without hardcoding.
        _gt_colors = {
            "shrouded":                        "#2196F3",
            "open_rotor":                      "#FF5722",
            "UAV / Drone":                     "#1565C0",
            "VTOL / Advanced Air Mobility":    "#E65100",
            "Other":                           "#9E9E9E",
            "unknown":                         "#9E9E9E",
        }
        _fallback = sns.color_palette("tab10", n_colors=len(gt_categories))
        gt_pal = {
            c: _gt_colors.get(c, _fallback[i % len(_fallback)])
            for i, c in enumerate(gt_categories)
        }

        gt_df = pd.DataFrame({
            "UMAP-1":    points_2d[:, 0],
            "UMAP-2":    points_2d[:, 1],
            "Class":     ground_truth_labels,
            "Patent ID": patent_ids,
        })
        sns.scatterplot(
            data=gt_df, x="UMAP-1", y="UMAP-2", hue="Class",
            palette=gt_pal, s=90, alpha=0.85,
            edgecolor="black", linewidth=0.4, ax=ax2,
        )
        ax2.set_title(
            f"Ground-Truth Labels{suffix}\n"
            f"(Shrouded vs Open Rotor)",
            fontsize=12, fontweight="bold",
        )
        ax2.set_xlabel("UMAP Dimension 1", fontsize=11)
        ax2.set_ylabel("UMAP Dimension 2", fontsize=11)
        ax2.legend(title="Class", bbox_to_anchor=(1.02, 1), loc="upper left", framealpha=0.9)
        ax2.grid(alpha=0.3)

    plt.tight_layout()
    return fig


def plot_cluster_gallery(
    cluster_df: pd.DataFrame,
    patent_ids: list[str],
    image_paths: list[Path],
    plot_dir: Path,
) -> None:
    """
    Save one image gallery PNG per non-noise cluster (4 columns, one image per patent).
    Noise points (cluster_id = −1) are excluded from the gallery.
    """
    plot_dir = Path(plot_dir)
    rep_image: dict[str, Path] = {}
    for p in image_paths:
        pid = patent_id_from_path(p)
        if pid not in rep_image:
            rep_image[pid] = p

    cluster_to_items: dict[int, list] = {}
    for pid in patent_ids:
        if pid not in rep_image:
            continue
        cid = int(cluster_df.loc[cluster_df["patent_id"] == pid, "cluster_id"].iloc[0])
        if cid >= 0:
            cluster_to_items.setdefault(cid, []).append((pid, rep_image[pid]))

    for cluster_id, items in sorted(cluster_to_items.items()):
        items = sorted(items, key=lambda x: x[0])
        n     = len(items)
        ncols = min(4, max(1, n))
        nrows = -(-n // ncols)  # ceiling division without math.ceil

        fig, axes = plt.subplots(nrows, ncols, figsize=(4 * ncols, 4 * nrows))
        if nrows == 1 and ncols == 1:
            axes = np.array([[axes]])
        elif nrows == 1:
            axes = axes[np.newaxis, :]
        elif ncols == 1:
            axes = axes[:, np.newaxis]

        for i in range(nrows * ncols):
            ax = axes[i // ncols, i % ncols]
            if i < n:
                pid, img_path = items[i]
                try:
                    ax.imshow(Image.open(img_path).convert("RGB"))
                    ax.set_title(pid, fontsize=9, fontweight="bold")
                except Exception:
                    ax.text(0.5, 0.5, f"Error\n{pid}", ha="center", va="center",
                            transform=ax.transAxes)
            ax.axis("off")

        fig.suptitle(f"Cluster {cluster_id} ({n} patents)", fontsize=14, fontweight="bold")
        fig.tight_layout()
        out = safe_save_plot(fig, plot_dir / f"cluster_{cluster_id}_gallery.png",
                             dpi=150, bbox_inches="tight")
        print(f"✓ Gallery saved: {out}")
        plt.show()
        plt.close(fig)


# ── KNN nearest-neighbor verification ────────────────────────────────────────

def plot_knn_neighbors(
    query_idx: int,
    image_paths: list[Path],
    image_emb: np.ndarray,
    k: int = 9,
) -> plt.Figure:
    """
    Display a query image alongside its K nearest neighbours in embedding space.

    Uses cosine similarity (L2-normalised dot product) to rank neighbours.
    The query is shown first with a green border; neighbours are ranked
    left-to-right, top-to-bottom by decreasing similarity.

    WHY THIS MATTERS FOR PATENT DATA:
    Nearest-neighbour inspection is the fastest way to diagnose what the
    embedding space has actually learned. If neighbours share the same drawing
    office style or background tone rather than structural geometry you know
    PCA / fine-tuning is needed before clustering results are trustworthy.

    Parameters
    ----------
    query_idx   : index into image_paths / image_emb to use as the query
    image_paths : list of image Paths aligned with image_emb rows
    image_emb   : (N, D) raw or L2-normalised embedding matrix
    k           : number of neighbours to display
    """
    X_unit    = normalize(image_emb, norm="l2")
    query_vec = X_unit[query_idx]
    sims      = X_unit @ query_vec
    sims[query_idx] = -2.0   # exclude self from ranking
    neighbor_indices = np.argsort(sims)[::-1][:k]

    n_panels = k + 1
    ncols    = min(n_panels, 5)
    nrows    = -(-n_panels // ncols)

    fig, axes = plt.subplots(nrows, ncols, figsize=(3 * ncols, 3.5 * nrows))
    axes = np.array(axes).flatten()

    # ── Query panel ───────────────────────────────────────────────────────
    ax = axes[0]
    try:
        ax.imshow(Image.open(image_paths[query_idx]).convert("RGB"))
    except Exception:
        ax.text(0.5, 0.5, "Error", ha="center", va="center", transform=ax.transAxes)
    for spine in ax.spines.values():
        spine.set_edgecolor("#43A047")
        spine.set_linewidth(3)
    ax.set_title(
        f"QUERY  (idx {query_idx})\n{image_paths[query_idx].name[:28]}",
        fontsize=7, color="#43A047", fontweight="bold",
    )
    ax.axis("off")

    # ── Neighbour panels ──────────────────────────────────────────────────
    for rank, n_idx in enumerate(neighbor_indices, start=1):
        if rank >= len(axes):
            break
        ax = axes[rank]
        try:
            ax.imshow(Image.open(image_paths[n_idx]).convert("RGB"))
        except Exception:
            ax.text(0.5, 0.5, "Error", ha="center", va="center", transform=ax.transAxes)
        ax.set_title(
            f"#{rank}  cos={sims[n_idx]:.3f}\n{image_paths[n_idx].name[:28]}",
            fontsize=6,
        )
        ax.axis("off")

    for ax in axes[n_panels:]:
        ax.axis("off")

    fig.suptitle(
        f"KNN Verification — Query idx {query_idx}  (top-{k} by cosine similarity)",
        fontsize=11, fontweight="bold",
    )
    plt.tight_layout()
    return fig


# ── Dimensionality-reduction comparison ──────────────────────────────────────

def plot_reduction_comparison(
    patent_emb: np.ndarray,
    patent_ids: list[str],
    hdbscan_min_cluster_size: int = 5,
    hdbscan_min_samples: int | None = None,
    pca_n_components: int = 100,
    seed: int = 42,
) -> plt.Figure:
    """
    Side-by-side: HDBSCAN on raw L2-normalised embeddings vs PCA-reduced.

    Both panels share the same UMAP 2D projection (computed once on the raw
    normalised embeddings) so spatial positions are comparable.  Only the
    colour assignment (which HDBSCAN run was used) differs.

    WHY THIS COMPARISON:
    Raw HDBSCAN in 1024-d suffers from the curse of dimensionality — distances
    are nearly uniform so density estimation is unreliable.  PCA reduces the
    space to directions of maximum variance.  If PCA produces tighter, more
    coherent clusters than raw clustering, it means the high-d space contained
    significant noise variance that PCA filtered out.  If results are similar,
    the top variance directions already dominate.

    A third panel shows the PCA scree curve so you can judge how much variance
    the chosen n_components captures.

    Parameters
    ----------
    patent_emb              : (N, D) raw patent-level embeddings
    patent_ids              : list of N patent ID strings
    hdbscan_min_cluster_size: shared between both HDBSCAN runs
    hdbscan_min_samples     : shared between both HDBSCAN runs
    pca_n_components        : components for the PCA-then-HDBSCAN branch
    seed                    : random seed for PCA and UMAP
    """
    # ── Normalise once ────────────────────────────────────────────────────
    X_norm = normalize(patent_emb, norm="l2")
    n_comp = min(pca_n_components, X_norm.shape[0], X_norm.shape[1])

    # ── UMAP projection (shared) ──────────────────────────────────────────
    print("Computing shared UMAP projection…")
    points_2d = umap_lib.UMAP(
        n_components=2, n_neighbors=15, min_dist=0.1,
        metric="euclidean", random_state=seed,
    ).fit_transform(X_norm)

    # ── Branch A: HDBSCAN on raw L2-normalised embeddings ─────────────────
    print(f"HDBSCAN on raw {X_norm.shape[1]}-d embeddings…")
    labels_raw = hdbscan_lib.HDBSCAN(
        min_cluster_size=hdbscan_min_cluster_size,
        min_samples=hdbscan_min_samples,
        metric="euclidean",
        cluster_selection_method="leaf",
    ).fit_predict(X_norm)
    n_raw   = len(set(labels_raw) - {-1})
    n_noise_raw = int(np.sum(labels_raw == -1))

    # ── Branch B: PCA then HDBSCAN ────────────────────────────────────────
    print(f"PCA {X_norm.shape[1]}d → {n_comp}d then HDBSCAN…")
    pca     = PCA(n_components=n_comp, random_state=seed)
    X_pca   = pca.fit_transform(X_norm)
    cumvar  = pca.explained_variance_ratio_.cumsum()
    var_cap = cumvar[-1] * 100

    labels_pca = hdbscan_lib.HDBSCAN(
        min_cluster_size=hdbscan_min_cluster_size,
        min_samples=hdbscan_min_samples,
        metric="euclidean",
        cluster_selection_method="leaf",
    ).fit_predict(X_pca)
    n_pca   = len(set(labels_pca) - {-1})
    n_noise_pca = int(np.sum(labels_pca == -1))

    print(f"Raw  : {n_raw} clusters, {n_noise_raw} noise")
    print(f"PCA  : {n_pca} clusters, {n_noise_pca} noise  ({var_cap:.1f}% variance in {n_comp}d)")

    # ── Plot ──────────────────────────────────────────────────────────────
    fig, axes = plt.subplots(1, 3, figsize=(21, 7))

    def _scatter(ax, labels, title):
        unique  = sorted(set(labels))
        n_real  = sum(1 for l in unique if l >= 0)
        palette = sns.color_palette("tab10", n_colors=max(n_real, 1))
        cmap, ci = {}, 0
        for lbl in unique:
            cmap[lbl] = (0.55, 0.55, 0.55) if lbl == -1 else palette[ci]
            if lbl >= 0:
                ci += 1
        colors = [cmap[l] for l in labels]
        ax.scatter(points_2d[:, 0], points_2d[:, 1],
                   c=colors, s=15, alpha=0.7, edgecolors="none")
        ax.set_title(title, fontsize=11, fontweight="bold")
        ax.set_xlabel("UMAP-1"); ax.set_ylabel("UMAP-2")
        ax.grid(alpha=0.2)
        # Legend patches
        from matplotlib.patches import Patch
        handles = [
            Patch(color=cmap[l], label="Noise" if l == -1 else f"C{l}")
            for l in unique[:12]   # cap at 12 to keep legend readable
        ]
        ax.legend(handles=handles, fontsize=7, ncol=2,
                  loc="upper right", framealpha=0.8)

    _scatter(axes[0], labels_raw,
             f"Raw L2-norm → HDBSCAN\n{n_raw} clusters  {n_noise_raw} noise")
    _scatter(axes[1], labels_pca,
             f"PCA-{n_comp}d → HDBSCAN\n{n_pca} clusters  {n_noise_pca} noise")

    # Scree panel
    ax3 = axes[2]
    ax3.plot(range(1, len(cumvar) + 1), cumvar * 100,
             linewidth=1.5, color="#1565C0")
    ax3.axhline(90, color="orange", linestyle="--", linewidth=1, label="90% var")
    ax3.axhline(95, color="red",    linestyle="--", linewidth=1, label="95% var")
    ax3.axvline(n_comp, color="green", linestyle=":", linewidth=1.5,
                label=f"PCA n={n_comp}  ({var_cap:.1f}%)")
    ax3.set_xlabel("PCA components"); ax3.set_ylabel("Cumulative variance (%)")
    ax3.set_title("PCA Scree Curve", fontsize=11, fontweight="bold")
    ax3.legend(fontsize=9); ax3.grid(alpha=0.3)

    fig.suptitle("Dimensionality Reduction Strategy Comparison\n"
                 "(UMAP projection shared — only cluster colours differ)",
                 fontsize=13, fontweight="bold")
    plt.tight_layout()
    return fig


# ── Excel analysis matrix export ──────────────────────────────────────────────

def export_analysis_matrix(
    image_paths: list[Path],
    dataset_df: pd.DataFrame,
    cluster_df: pd.DataFrame,
    output_path: Path,
) -> pd.DataFrame:
    """
    Compile and export a structured Excel spreadsheet for offline analysis.

    Joins per-image metadata (from DeepPatent2Dataset.build_dataframe()) with
    per-patent cluster assignments (from hdbscan_cluster()), and saves the
    result as a formatted .xlsx file.

    Columns in the output sheet
    ---------------------------
    Image_ID        : filename without extension
    File_Path       : absolute path to the source image
    Patent_ID       : e.g. USD0850055-20190604
    Year            : 4-digit year (or year-range key like 2019part3)
    Object_Title    : most common object label across subfigures on this page
    Aspect_Views    : semicolon-separated list of unique viewing angles
    BBox_Count      : number of subfigure annotations for this image page
    Caption_Sample  : first caption text from the JSON annotations
    Cluster_ID      : HDBSCAN cluster assignment (-1 = noise)
    Cluster_Prob    : HDBSCAN membership probability (0–1; absent for noise)

    Parameters
    ----------
    image_paths : list of image Paths (same order used during embedding)
    dataset_df  : DataFrame from DeepPatent2Dataset.build_dataframe()
    cluster_df  : DataFrame from hdbscan_cluster() [patent_id, cluster_id, cluster_prob]
    output_path : where to write deeppatent_analysis_matrix.xlsx
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── Build base table from dataset_df ─────────────────────────────────
    # dataset_df is indexed by figure_file; ensure we only include images
    # that were actually processed (subset / sampled by the pipeline).
    processed_names = {p.name for p in image_paths}
    base = dataset_df[dataset_df["figure_file"].isin(processed_names)].copy()

    base = base.rename(columns={
        "figure_file":    "Image_ID",
        "file_path":      "File_Path",
        "patentID":       "Patent_ID",
        "year":           "Year",
        "platform":       "Platform",
        "object_title":   "Object_Title",
        "aspects":        "Aspect_Views",
        "bbox_count":     "BBox_Count",
        "caption_sample": "Caption_Sample",
    })
    base["Image_ID"] = base["Image_ID"].str.replace(r"\.png$", "", regex=True)

    # ── Join cluster assignments (patent level → image level) ─────────────
    cluster_lookup = cluster_df.set_index("patent_id")[["cluster_id", "cluster_prob"]]
    base = base.join(
        cluster_lookup.rename(columns={
            "cluster_id":   "Cluster_ID",
            "cluster_prob": "Cluster_Prob",
        }),
        on="Patent_ID",
        how="left",
    )
    base["Cluster_ID"]   = base["Cluster_ID"].fillna(-1).astype(int)
    base["Cluster_Prob"] = base["Cluster_Prob"].fillna(0.0).round(4)

    # ── Column order ──────────────────────────────────────────────────────
    cols = [
        "Image_ID", "File_Path", "Patent_ID", "Year", "Platform",
        "Object_Title", "Aspect_Views", "BBox_Count", "Caption_Sample",
        "Cluster_ID", "Cluster_Prob",
    ]
    base = base[[c for c in cols if c in base.columns]].reset_index(drop=True)

    # ── Write Excel with basic formatting ─────────────────────────────────
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        base.to_excel(writer, sheet_name="Analysis", index=False)

        ws = writer.sheets["Analysis"]

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-fit column widths (capped at 60)
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value),
                default=10,
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0",
                                  fill_type="solid")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

    n_clusters = base["Cluster_ID"].nunique()
    n_noise    = int((base["Cluster_ID"] == -1).sum())
    print(f"[export_analysis_matrix] {len(base)} rows | "
          f"{n_clusters} unique cluster IDs | {n_noise} noise images")
    print(f"Saved → {output_path}")
    return base
