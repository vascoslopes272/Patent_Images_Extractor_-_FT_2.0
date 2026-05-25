"""
patents.py — Load the Excel and return patent subsets.

Usage in any notebook:
    from src.patents import load_patents, get_subset
    df = load_patents(cfg)
    subset = get_subset(df, cfg)
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path


def load_patents(cfg: dict) -> pd.DataFrame:
    """
    Load the Excel and attach the real PDF URLs (extracted from hyperlinks).
    Returns a clean DataFrame with a 'pdf_url' column.
    """
    excel_path = Path(cfg["paths"]["excel"])
    print(f"Loading Excel: {excel_path}")

    # 1. Load data with pandas (fast, gets all column values)
    df = pd.read_excel(excel_path, dtype={"Record Number": str})

    # 2. Extract hyperlink URLs from the PDF Link column using openpyxl
    wb = load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    pdf_col_idx = headers.index("PDF Link") + 1  # 1-based

    urls = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=pdf_col_idx, max_col=pdf_col_idx):
        cell = row[0]
        if cell.hyperlink and cell.hyperlink.target:
            urls.append(cell.hyperlink.target)
        else:
            urls.append(None)

    # Align length with df (Excel may have trailing empty rows)
    urls = urls[: len(df)]
    while len(urls) < len(df):
        urls.append(None)

    df["pdf_url"] = urls
    df = df[df["pdf_url"].notna()].copy()
    df = df.reset_index(drop=True)
    print(f"  Loaded {len(df)} patents with valid PDF URLs.")
    return df


def get_subset(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """
    Return a subset of patents based on config['subset'] settings.

    Modes:
      'all'      → all patents
      'n_first'  → first N patents
      'filter'   → filter by record_type, legal_status, tech_sub_domain
    """
    mode = cfg["subset"]["mode"]

    if mode == "all":
        subset = df.copy()

    elif mode == "n_first":
        n = cfg["subset"]["n_first"]
        subset = df.head(n).copy()

    elif mode == "filter":
        subset = df.copy()
        filters = cfg["subset"]["filters"]

        if filters.get("record_type"):
            subset = subset[subset["Record Type"] == filters["record_type"]]

        if filters.get("legal_status"):
            subset = subset[
                subset["Family Legal Status(Dead/Alive)"] == filters["legal_status"]
            ]

        if filters.get("tech_sub_domain"):
            # partial match — "TRANSPORT" matches "TRANSPORT, CONTROL"
            subset = subset[
                subset["Tech Sub Domain"]
                .str.contains(filters["tech_sub_domain"], na=False)
            ]

    else:
        raise ValueError(f"Unknown subset mode: '{mode}'. Use 'all', 'n_first', or 'filter'.")

    subset = subset.reset_index(drop=True)
    print(f"  Subset mode='{mode}': {len(subset)} patents selected.")
    return subset
