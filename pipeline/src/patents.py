"""
patents.py — Load the Excel dataset and select the patent subset to process.

Two functions used by every notebook and main.py:
    df, missing_df = load_patents(cfg)   # reads Excel, extracts real PDF URLs
    subset         = get_subset(df, cfg) # applies the subset filter from config.yaml

load_patents() returns TWO DataFrames:
    df         — patents that have a valid PDF URL (ready to process)
    missing_df — patents with no URL at all (pass to extract_crops_streaming
                 so they appear in the combined extraction_status.xlsx report)
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path


# ---------------------------------------------------------------------------
# load_patents
# ---------------------------------------------------------------------------

def load_patents(cfg: dict):
    """
    Open the Excel, extract real PDF hyperlink URLs, return two DataFrames.

    WHY TWO LIBRARIES?
    pandas.read_excel reads cell values (text, numbers) quickly but cannot see
    hyperlinks embedded in cells — it only sees the display text, not the URL.
    openpyxl opens the raw XML of the .xlsx file and can read the actual hyperlink
    target (the patseer API URL) stored inside the "PDF Link" column cells.

    RETURNS:
        (df_valid, df_missing)
        df_valid   — rows with a usable pdf_url  → pass to get_subset() then extract
        df_missing — rows with no pdf_url at all → pass to extract_crops_streaming()
                     so they are recorded as "no_url" in extraction_status.xlsx

    FILES READ:
        Excel at cfg["paths"]["excel"]   (read twice: once by pandas, once by openpyxl)

    FILES WRITTEN:
        none — the combined status report is written by extract_crops_streaming()
    """
    excel_path = Path(cfg["paths"]["excel"])
    print(f"Loading Excel: {excel_path}")

    # ── Pass 1: pandas reads all cell values (fast) ───────────────────────
    df = pd.read_excel(excel_path, dtype={"Record Number": str})

    # ── Pass 2: openpyxl extracts the hyperlink URLs ──────────────────────
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find which column "PDF Link" is in (1-based index for openpyxl)
    headers = [cell.value for cell in ws[1]]
    pdf_col_idx = headers.index("PDF Link") + 1

    # Walk every data row and grab the hyperlink target if it exists
    urls = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                             min_col=pdf_col_idx, max_col=pdf_col_idx):
        cell = row[0]
        if cell.hyperlink and cell.hyperlink.target:
            urls.append(cell.hyperlink.target)   # real patseer API URL
        else:
            urls.append(None)                    # no hyperlink in this cell

    # openpyxl may return more rows than pandas if the sheet has trailing blank
    # rows — clip to match the DataFrame length, then pad if shorter
    urls = urls[: len(df)]
    while len(urls) < len(df):
        urls.append(None)

    df["pdf_url"] = urls

    # ── Split into valid and missing ──────────────────────────────────────
    missing = df[df["pdf_url"].isna()].copy().reset_index(drop=True)
    valid   = df[df["pdf_url"].notna()].copy().reset_index(drop=True)

    if not missing.empty:
        print(f"  {len(missing)} patents have no PDF URL (will appear in status report)")
    print(f"  Loaded {len(valid)} patents with valid PDF URLs.")
    return valid, missing


# ---------------------------------------------------------------------------
# get_subset
# ---------------------------------------------------------------------------

def get_subset(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """
    Slice the full DataFrame down to the patents you actually want to process.

    The selection mode is controlled by config.yaml → subset.mode:

    "all"     → use every patent that has a valid URL (all 1618)

    "n_first" → take the first N rows from the top of the DataFrame.
                N is set by config.yaml → subset.n_first.
                Good for quick tests (e.g. n_first: 5) before running the full dataset.

    "filter"  → keep only rows that match ALL non-null filter values in
                config.yaml → subset.filters:
                  record_type      e.g. "Patent" (exact match)
                  legal_status     e.g. "ALIVE"  (exact match)
                  tech_sub_domain  e.g. "TRANSPORT" (partial match — a row with
                                   "TRANSPORT, CONTROL" would be included)
                  cpc_first        list of CPC codes — keeps only rows whose FIRST
                                   CPC code matches one of these exactly (the CPC
                                   column stores codes separated by " | ")

    "selected" → keep only the patents manually flagged in the selector UI.
                 Reads logs/selected_patents.json (written by 01_patent_selector.ipynb).
                 Run 01_patent_selector.ipynb first; then set mode: "selected" here.

    RETURNS: a new DataFrame (copy), rows reset to 0-based index.
    """
    mode = cfg["subset"]["mode"]

    if mode == "all":
        # Process every patent that has a valid URL
        subset = df.copy()

    elif mode == "n_first":
        # Take the first N rows — useful for testing before running on all 1618
        n = cfg["subset"]["n_first"]
        subset = df.head(n).copy()

    elif mode == "filter":
        subset = df.copy()
        filters = cfg["subset"]["filters"]

        if filters.get("record_type"):
            # Exact match on the "Record Type" column (e.g. "Patent")
            subset = subset[subset["Record Type"] == filters["record_type"]]

        if filters.get("legal_status"):
            # Exact match on Family Legal Status (e.g. "ALIVE" or "DEAD")
            subset = subset[
                subset["Family Legal Status(Dead/Alive)"] == filters["legal_status"]
            ]

        if filters.get("tech_sub_domain"):
            # Partial match — "TRANSPORT" matches "TRANSPORT, CONTROL" etc.
            subset = subset[
                subset["Tech Sub Domain"]
                .str.contains(filters["tech_sub_domain"], na=False)
            ]

        if filters.get("cpc_first"):
            # Keep only rows where the FIRST CPC code matches one of the listed values.
            # The CPC column stores codes separated by " | "; we compare only the first one.
            allowed = [c.strip() for c in filters["cpc_first"]]
            first_cpc = df["CPC"].fillna("").str.split(r"\s*\|\s*", n=1).str[0].str.strip()
            subset = subset[first_cpc.isin(allowed)]

    elif mode == "selected":
        import json
        sel_path = Path(cfg["paths"]["logs"]) / "selected_patents.json"
        if not sel_path.exists():
            raise FileNotFoundError(
                f"selected_patents.json not found at {sel_path}.\n"
                "Run 01_patent_selector.ipynb first and mark the patents you want."
            )
        with open(sel_path) as f:
            data = json.load(f)
        selected_ids = {
            str(pid).strip().upper()
            for pid, v in data.get("patents", {}).items()
            if v.get("selected")
        }
        subset = df[df["Record Number"].apply(
            lambda x: str(x).strip().upper() in selected_ids
        )].copy()

    else:
        raise ValueError(f"Unknown subset mode: '{mode}'. Use 'all', 'n_first', 'filter', or 'selected'.")

    subset = subset.reset_index(drop=True)
    print(f"  Subset mode='{mode}': {len(subset)} patents selected.")
    return subset
